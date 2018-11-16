#!/usr/bin/env python3
#
# enumsheets.py
#
# Counts total number of dxf files specified on the command line. Writes sheet
# number and total number of sheets to every dxf files' title block. Optionally
# rewrites date, scale, and address fields. Extracts each sheet title and saves
# contents into Excel table.
#
# This script probably is not suited for generic use without additional
# configuration or even modifications. It depends on many custom aspects
# used in production of drawings for the needs of artidea.gallery's design
# projects.
#
# This script is only tested to work with dxf files created and edited with
# LibreCAD (https://librecad.org). Never tested it with files created using 
# another CAD software.
#
# In order to be counted and processed a dxf file must contain a title block
# which corresponds to the following requirements:
#
# * it must be an INSERT in dxf file;
# * it must contain a marker as defined by
#     conf['title_block']['marker'] in the code below;
#
# If such a block is found, this script tries to find corresponding title block
# fields. If title block is just inserted from template, those fields would
# contain the following ready for replacement placeholder markers:
#
# X - for the 'Sheet number' field;
# XX - for the 'Number of sheets' field;
# TitleField - for the 'Sheet title' field;
# AddressField - for the 'Address' field;
# 1:50 - for the 'Scale' field;
# 0000-00-00 - for the 'Sheet date' field.
#
# In my workflow, during sheet editing the 'TitleField' is usually manually
# replaced by some text starting from words 'План' or 'Развёртка' (in Russian).
# Yours is obviously different. You can adjust corresponding regular expression
# in configuration file. Also you probably need to adjust regular expression
# for the 'Address' field.
#
# This script should be able to process not only freshly inserted from
# template title block, but also modified title block whose field markers
# are already replaced with real values. As far as 'Sheet number' and
# 'Number of sheets' may contain values of the same pattern, some dumb logic
# is used to guess which value belongs to what field.
#
# Copyright (C) 2018 Alexander Pravdin <aledin@mail.ru>
#
# License: MIT
#

import os
import re
import sys
import time
import ezdxf
import argparse
import configparser


# All default values can be overriden in configuration file.

DEFAULT_TITLE_BLOCK_MARKER = 'artidea.gallery'

DEFAULT_OUTPUT_DIRNAME = 'enumerated_sheets'

# 'Sheet number' field should contain either 'X' or a number.
DEFAULT_FLD_NUMBER_PATTERN = r'^(X|\d{1,3})$'

# 'Number of sheets' field should contain either 'XX' or a number.
DEFAULT_FLD_SHEETS_PATTERN = r'^(XX|\d{1,3})$'

# 'Sheet title' field should contain either 'TitleField'
# or a line of text starting with 'План' or 'Разв' (in Russian).
# Yours probably is different. Adjust your pattern in configuration file.
DEFAULT_FLD_TITLE_PATTERN = r'((^TitleField$)|(^(План|Разв)))'

# 'Address' field should contain either 'AddressField' or a line
# of text starting with 'г.' (Russian small 'Г' letter). Your pattern
# probably is different. Adjust it in configuration file. 
DEFAULT_FLD_ADDRESS_PATTERN = r'((^AddressField$)|(^г))'

# 'Sheet date' field should contain either '0000-00-00'
# or a date like '2018-11-24'.
DEFAULT_FLD_DATE_PATTERN = r'^(\d{4}-\d{2}-\d{2})$'

# 'Scale' field.
DEFAULT_FLD_SCALE_PATTERN = r'^(\d+:\d+)$'

# Update date field?
DEFAULT_UPDATE_DATE = True

# Update scale field?
DEFAULT_UPDATE_SCALE = True

# Update address field?
DEFAULT_UPDATE_ADDRESS = True

# Save to Excel file?
DEFAULT_EXCEL_ENABLE = True

DEFAULT_EXCEL_FILENAME = 'contents.xlsx'
DEFAULT_EXCEL_WORKSHEET_TITLE = 'Перечень листов'
DEFAULT_EXCEL_DRAWINGS_TITLE = 'Чертежи'
DEFAULT_EXCEL_SPECS_TITLE = 'Ведомости'


conf = {
        'output': {
            'dirname': DEFAULT_OUTPUT_DIRNAME,
        },
        'title_block': {
            'marker': DEFAULT_TITLE_BLOCK_MARKER,
            'fields': {
                'number': {
                    're': re.compile(DEFAULT_FLD_NUMBER_PATTERN),
                },
                'date': {
                    're': re.compile(DEFAULT_FLD_DATE_PATTERN),
                    'value': '',
                },
                'title': {
                    're': re.compile(DEFAULT_FLD_TITLE_PATTERN),
                },
                'sheets': {
                    're': re.compile(DEFAULT_FLD_SHEETS_PATTERN),
                },
                'address': {
                    're': re.compile(DEFAULT_FLD_ADDRESS_PATTERN),
                    'value': '',
                },
                'scale': {
                    're': re.compile(DEFAULT_FLD_SCALE_PATTERN),
                    'value': '',
                },
            },
            'update_date': DEFAULT_UPDATE_DATE,
            'update_scale': DEFAULT_UPDATE_SCALE,
            'update_address': DEFAULT_UPDATE_ADDRESS,
        },
        'excel_file': {
            'enable': DEFAULT_EXCEL_ENABLE,
            'filename': DEFAULT_EXCEL_FILENAME,
            'worksheet_title': DEFAULT_EXCEL_WORKSHEET_TITLE,
            'drawings_title': DEFAULT_EXCEL_DRAWINGS_TITLE,
            'specs_title': DEFAULT_EXCEL_SPECS_TITLE,
            'specs_names': tuple(),
        },
}


def parse_config(config_f):
    ini = configparser.ConfigParser()
    ini.read_file(f)

    conf['output']['dirname'] = ini.get(
            'output', 'dirname',
            fallback=DEFAULT_OUTPUT_DIRNAME)

    conf['title_block']['marker'] = ini.get(
            'title_block', 'marker',
            fallback=DEFAULT_TITLE_BLOCK_MARKER)

    patpar = (
            ('number',   DEFAULT_FLD_NUMBER_PATTERN),
            ('date',     DEFAULT_FLD_DATE_PATTERN),
            ('title',    DEFAULT_FLD_TITLE_PATTERN),
            ('sheets',   DEFAULT_FLD_SHEETS_PATTERN),
            ('address',  DEFAULT_FLD_ADDRESS_PATTERN),
            ('scale',    DEFAULT_FLD_SCALE_PATTERN),
    )

    for p, fb in patpar:
        conf['title_block']['fields'][p]['re'] = re.compile(ini.get(
            'title_block', p + '_pattern', fallback=fb))

    updpar = (
            ('date',     DEFAULT_UPDATE_DATE),
            ('scale',    DEFAULT_UPDATE_SCALE),
            ('address',  DEFAULT_UPDATE_ADDRESS),
    )

    for p, fb in updpar:
        key = 'update_' + p
        conf['title_block'][key] = ini.getboolean(
                'title_block', key, fallback=fb)

    conf['title_block']['fields']['date']['value'] = ini.get(
            'title_block', 'date_value', fallback='')
    conf['title_block']['fields']['scale']['value'] = ini.get(
            'title_block', 'scale_value', fallback='')
    conf['title_block']['fields']['address']['value'] = ini.get(
            'title_block', 'address_value', fallback=''
            ).strip().replace('\n', '\P')

    conf['excel_file']['enable'] = ini.getboolean(
            'excel_file', 'enable',
            fallback=DEFAULT_EXCEL_ENABLE)
    conf['excel_file']['filename'] = ini.get(
            'excel_file', 'filename',
            fallback=DEFAULT_EXCEL_FILENAME)
    conf['excel_file']['worksheet_title'] = ini.get(
            'excel_file', 'worksheet_title',
            fallback=DEFAULT_EXCEL_WORKSHEET_TITLE)
    conf['excel_file']['drawings_title'] = ini.get(
            'excel_file', 'drawings_title',
            fallback=DEFAULT_EXCEL_DRAWINGS_TITLE)
    conf['excel_file']['specs_title'] = ini.get(
            'excel_file', 'specs_title',
            fallback=DEFAULT_EXCEL_SPECS_TITLE)
    specs_names = tuple(x for x in ini.get(
        'excel_file', 'specs_names', fallback='').split('\n') if x.strip())
    conf['excel_file']['specs_names'] = specs_names


def __make_output_dir(dirname, ext_cnt=0):
    '''
    Makes output directory. If directory exists, create another one
    with the same name and appended number as an extension.
    
    @dirname: directory name to create
    @ext_cnt: iteration counter for searching unused extension number

    Returns created dirname on success, or '' if too many dirs with the same
    names and num extensions already exist.
    '''
    if ext_cnt >= 1000:  # too many dirs with the same names and num extensions
        return ''
    try:
        os.mkdir(dirname)
    except FileExistsError:
        ext_cnt += 1
        dirname = "{:s}.{:03d}".format(os.path.splitext(dirname)[0], ext_cnt)
        return __make_output_dir(dirname, ext_cnt)
    return dirname


def find_our_title_block(dwg):
    '''
    Searches for a block which contains our unique marker.
    Such a block is our title block in which we'll write our values.

    @dwg: drawing
    
    Returns query result (reference to the found block).
    '''
    for b in dwg.blocks:
        res = b.query('MTEXT')
        for t in res:
            if conf['title_block']['marker'] in t.get_text():
                return res
    return None


def recognize_our_sheets(dxf_files):
    '''
    Finds out which of the dxf files are our drawing sheets.

    @dxf_files: a list of dxf files

    Returns a tuple of our_sheets and other_files.
        our_sheets - a list of DrawingSheet objects (our recognized sheets)
        other_files - a list of filenames of unrecognized files.
    '''
    our_sheets = []
    other_files = []
    for f in dxf_files:
        print("Looking for title block in %s: " % (f), end='', flush=True)
        dwg = ezdxf.readfile(f)
        res = find_our_title_block(dwg)
        if res:
            print("Found, will process this file")
            our_sheets.append(DrawingSheet(dwg, res))
        else:
            print("Not found, will copy this file 'as is'")
            other_files.append(f)
    return (our_sheets, other_files)


def enumerate_sheets(sheets, output_dir):
    '''
    Sorts sheets by filename, calculates number of sheets, sets sheets values,
    saves processed sheets to @output_dir directory.

    @sheets: a list of DrawingSheet objects
    @output_dir: directory where to save processed dxf files
    '''
    sheets.sort(key=lambda x: x.dwg.filename)
    num_of_sheets = len(sheets)
    sheet_num = 0
    update_date = conf['title_block']['update_date']
    update_scale = conf['title_block']['update_scale']
    update_address = conf['title_block']['update_address']
    new_date = conf['title_block']['fields']['date']['value']
    new_scale = conf['title_block']['fields']['scale']['value']
    new_address = conf['title_block']['fields']['address']['value']

    for x in sheets:
        output_file = os.path.join(output_dir, os.path.basename(x.dwg.filename))
        print("Processing %s, saving to %s..." %
                (x.dwg.filename, output_file))
        sheet_num += 1
        x.set_numbers(sheet_num, num_of_sheets)
        if update_date:
            if not new_date:
                new_date = time.strftime('%Y-%m-%d', time.localtime())
            x.set_date(new_date)
        if update_scale:
            if new_scale:
                scale = new_scale
            else:
                scale = x.get_dwg_scale()
            x.set_scale(scale)
        if update_address:
            if new_address:
                x.set_address(new_address)
        x.dwg.saveas(output_file)


def save_to_excel_table(drawing_sheets, output_dir):
    '''
    Creates contents page in Excel file.

    @drawing_sheets: a list of DrawingSheet objects
    @output_dir: directory where to save Excel file

    FIXME: Cannot work correctly with number of sheets
           exceeding of what can fit to an A3 page
           in two columns.
    '''

    if not conf['excel_file']['enable']:
        return

    filename = conf['excel_file']['filename']

    if not filename.strip():
        return

    ws_title = conf['excel_file']['worksheet_title']
    drawings_title = conf['excel_file']['drawings_title']
    specs_title = conf['excel_file']['specs_title']
    specs_names = conf['excel_file']['specs_names']
 
    output_file = os.path.join(output_dir, filename)
    print("Saving contents to %s..." % (output_file))

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()

    ws = wb.active
    ws.title = ws_title

    ws['A2'] = drawings_title
    ws['A2'].font = Font(name='Liberation Sans', size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:B2')

    max_rows_per_page = 47  # FIXME: should be configurable

    row_idx = 4
    row_offset = 0
    col_offset = 0

    for x in drawing_sheets:

        list_num = x.get_number()
        list_title = x.get_title().replace(r'\P', ' ')

        print('%s: %s' % (list_num, list_title))

        if row_idx > max_rows_per_page:
            col_offset = 3
            row_offset = max_rows_per_page - 3

        cell = ws.cell(
                row=row_idx-row_offset,
                column=1+col_offset,
                value=list_num)
        cell.font = Font(name='Liberation Sans', size=12)
        cell.alignment = Alignment(horizontal='center')

        cell = ws.cell(
                row=row_idx-row_offset,
                column=2+col_offset,
                value=list_title)
        cell.font = Font(name='Liberation Sans', size=12)
        #cell.alignment = Alignment(shrink_to_fit=True)

        row_idx += 1

    if specs_names: # Additional worksheet contents

        col_offset = 3

        if row_idx <= max_rows_per_page:
            row_idx = 2
            row_offset = 0
        else:
            row_idx += 2

        cell = ws.cell(
            row=row_idx-row_offset,
            column=1+col_offset,
            value=specs_title)
        cell.font = Font(name='Liberation Sans', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')

        ws.merge_cells(
            start_row=row_idx-row_offset,
            end_row=row_idx-row_offset,
            start_column=1+col_offset,
            end_column=2+col_offset)

        row_idx += 2

        for v in specs_names:
            cell = ws.cell(
                row=row_idx-row_offset,
                column=2+col_offset,
                value=v)
            cell.font = Font(name='Liberation Sans', size=12)
            row_idx += 1

    wb.save(output_file)


class DrawingSheet:

    def __init__(self, dwg, qres):
        self.dwg = dwg
        self.qres = qres

        # Title block fields
        # idx: field index
        # func: this function will be used to reduce
        #       multiple found indexes to only one.
        self.tb_fields = {
            'number':  { 'idx': None, 'func': min, },
            'sheets':  { 'idx': None, 'func': max, },
            'date':    { 'idx': None, 'func': lambda x: x[0], },
            'scale':   { 'idx': None, 'func': lambda x: x[0], },
            'title':   { 'idx': None, 'func': max, },
            'address': { 'idx': None, 'func': min, },
        }

        for name, data in self.tb_fields.items():
            self.__build_get_set_funcs(name)
            re = conf['title_block']['fields'][name]['re']
            func = data['func']
            data['idx'] = self.__find_fld_idx(re, func)

    def __build_get_set_funcs(self, name):
        '''
        Dynamically builds get_* and set_* funcions
        '''

        def getter():
            idx = self.tb_fields[name]['idx']
            if idx is None:
                return ''
            if self.qres is None:
                return ''
            return self.qres[idx].get_text()

        def setter(value):
            idx = self.tb_fields[name]['idx']
            if idx is None:
                return
            if self.qres is None:
                return
            self.qres[idx].set_text(str(value))

        setattr(self, 'get_' + name, getter)
        setattr(self, 'set_' + name, setter)

    def __find_fld_idx(self, re, func):
        '''
        Searches for all fields indexes matching regular expression @re.
        Applies function @func to return only one of them.
        
        @re:  compiled regular expression for the search
        @func: function to apply to multiple found indexes in order
               to choose only one

        Returns value depending on @func, or None if nothing is found.
        '''

        idx_list = self.__find_matching_indexes(re)

        if idx_list:
            return func(idx_list)


    def __find_matching_indexes(self, re):
        '''
        Finds all indexes matching critera.

        @re: compiled regular expression criterium for the search

        Returns a list of found indexes.
        '''
        res = []
        for i in range(len(self.qres)):
            t = self.qres[i].get_text()
            if re.search(t):
                res.append(i)
        return res

    def set_numbers(self, sheet_num, num_of_sheets):
        
        if self.qres is None:
            return

        # First, get values already present in title block
        try:
            old_number = int(self.get_number())
        except ValueError:
            old_number = 0
       
        try:
            old_sheets = int(self.get_sheets())
        except ValueError:
            old_sheets = 0

        if old_number > old_sheets:
            # Obviously our index findings were wrong, swap indexes
            self.tb_fields['sheets']['idx'], self.tb_fields['number']['idx'] = \
            self.tb_fields['number']['idx'], self.tb_fields['sheets']['idx']

        self.set_number(sheet_num)
        self.set_sheets(num_of_sheets)

    def get_dwg_scale(self):
        '''
        Returns text representation for scale set for drawing file itself.
        '''
        try:
            psvpscale = float(self.dwg.header.get('$PSVPSCALE', 1.0))
        except ValueError:
            psvpscale = 1.0
        if psvpscale > 1.0:
            scale = str(round(psvpscale)) + ':1'
        elif psvpscale < 1.0:
            scale = '1:' + str(round(1/psvpscale))
        else:
            scale = '1:1'
        return scale


if __name__ == '__main__':

    argparser = argparse.ArgumentParser()

    argparser.add_argument("-c", "--config",
            help="Configuration ini file",
            default=os.path.join(os.path.dirname(__file__), 'config.ini'))
    argparser.add_argument("dxf_file", nargs='+')

    args = argparser.parse_args()

    try:
        with open(args.config, encoding='utf-8') as f:
            parse_config(f)
    except Exception as e:
        print(e)
        sys.exit(1)

    #import pprint
    #pprint.pprint(conf)
    #sys.exit()

    #if len(sys.argv) < 2:
    #    argparser.print_usage()
    #    sys.exit(1)

    #dxf_files = [f for f in sys.argv if os.path.splitext(f)[1] == '.dxf']
    dxf_files = [f for f in args.dxf_file if os.path.splitext(f)[1] == '.dxf']

    (our_sheets, other_files) = recognize_our_sheets(dxf_files)

    if not (our_sheets or other_files):
        print('No input files to process.')
        sys.exit(1)

    output_dir = __make_output_dir(conf['output']['dirname'])

    if '' == output_dir:
        print('Error: Cannot create output directory.')
        sys.exit(1)

    if other_files:
        import shutil
        for f in other_files:
            print("Copying %s to %s..." %
                    (f, os.path.join(output_dir, os.path.basename(f))))
            shutil.copy(f, output_dir)

    if not our_sheets:
        print('No sheets to process.')
        sys.exit(1)
 
    # Count sheets, sort them, write numbers, extract titles.
    enumerate_sheets(our_sheets, output_dir)

    # Write contents to Excel table.
    save_to_excel_table(our_sheets, output_dir)

    print('Done.')

